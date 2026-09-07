import streamlit as st
import json
import zipfile
import time
import re
import io
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================
# PAGE CONFIG
# =========================
st.set_page_config(
    page_title="OEM Multi-Studio PRO",
    page_icon="🚀",
    layout="wide"
)

# =========================
# SIDEBAR NAVIGATION (Slide Selector)
# =========================
st.sidebar.title("Navigation")
app_mode = st.sidebar.radio(
    "Choose Tool / Work Mode:",
    ["🚀 Json Merged And Clean Data", "📊 Model Comparison"]
)

# =========================
# MODE 1: Json Merged And Clean Data
# =========================
if app_mode == "🚀 Json Merged And Clean Data":

    # CONFIG & MAPS
    REMOVE_VALUES = {"no", "not applicable", "not available", "0"}
    REMOVE_EXTRA = {"available", "standard", "optional", "option", "0"}
    BLANK_HEADERS = {"features", "options", "videos", "attachments"}

    NUMBER_MAP = {
        "0": "Zero", "1": "One", "2": "Two", "3": "Three",
        "4": "Four", "5": "Five", "6": "Six", "7": "Seven",
        "8": "Eight", "9": "Nine"
    }

    INDEX_WORD_MAP = {
        2: "Two", 3: "Three", 4: "Four", 5: "Five", 
        6: "Six", 7: "Seven", 8: "Eight", 9: "Nine", 10: "Ten"
    }

    # --- CAMELCASE HELPERS ---
    def number_to_word(num):
        return ''.join(NUMBER_MAP[d] for d in num if d in NUMBER_MAP)

    def to_camel_case(text, keep_number=False):
        text = re.sub(r'[^a-zA-Z0-9]+', ' ', text)
        words = re.findall(r'[A-Za-z0-9]+', text)
        cleaned = []
        
        for word in words:
            if word.isdigit():
                if keep_number:
                    for digit in word:
                        if digit in NUMBER_MAP:
                            cleaned.append(NUMBER_MAP[digit])
                continue
            
            if re.search(r'[a-zA-Z]', word) and re.search(r'\d', word):
                sub_tokens = re.findall(r'[a-zA-Z]+|\d+', word)
                for token in sub_tokens:
                    if token.isdigit():
                        if keep_number:
                            for digit in token:
                                if digit in NUMBER_MAP:
                                    cleaned.append(NUMBER_MAP[digit])
                    else:
                        cleaned.append(token)
            else:
                cleaned.append(word)
                
        if not cleaned:
            return ""
            
        return cleaned[0].lower() + ''.join(w.capitalize() for w in cleaned[1:])

    def get_base_label(label):
        return re.sub(r'\d+', '', label).strip().lower()
    
    # --- DYNAMIC CAMELCASE PROCESSOR ---
    def process_json_keys_to_camel(data):
        if isinstance(data, dict):
            label_count = {}
            for value in data.values():
                if isinstance(value, dict) and "label" in value:
                    base = get_base_label(value["label"])
                    label_count[base] = label_count.get(base, 0) + 1

            new_dict = {}
            for key, value in data.items():
                if isinstance(value, dict) and "label" in value and "desc" in value:
                    label = value["label"]
                    base = get_base_label(label)
                    if label_count.get(base, 0) > 1:
                        correct_key = to_camel_case(label, keep_number=True)
                    else:
                        correct_key = to_camel_case(label, keep_number=False)
                else:
                    correct_key = key

                if isinstance(value, dict):
                    value = process_json_keys_to_camel(value)
                elif isinstance(value, list):
                    value = [process_json_keys_to_camel(i) for i in value]

                final_key = correct_key
                count = 2
                while final_key in new_dict:
                    suffix_word = INDEX_WORD_MAP.get(count, f"Copy{count}")
                    final_key = f"{correct_key}{suffix_word}"
                    count += 1

                new_dict[final_key] = value
            return new_dict
        elif isinstance(data, list):
            return [process_json_keys_to_camel(i) for i in data]
        return data

    # --- CLEANING ENGINES ---
    def clean_string_global(s):
        if not isinstance(s, str): return s
        s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ").replace("\xa0", " ")
        return re.sub(r"\s+", " ", s).strip()

    def clean_dots_spec_string(text):
        if not isinstance(text, str): return text
        protected = {}
        def protect_decimal(m):
            key = f"__DEC_{len(protected)}__"
            protected[key] = m.group(0)
            return key
        text = re.sub(r'\d+\.\d+|\.\d+', protect_decimal, text)
        text = re.sub(r'(?i)(?<=[a-z])\.', ' ', text)
        for k, v in protected.items():
            text = text.replace(k, v)
        return clean_string_global(text)

    def normalize_text(text):
        text = clean_string_global(text)
        text = re.sub(r'^(?:\d+\s+)', '', text)
        text = re.sub(r'(?:\s+\d+)$', '', text)
        return text.lower().strip()

    def is_blank_header(key, value):
        return key.lower() in BLANK_HEADERS and value in ("", None, [], {})

    def is_spec_object(obj):
        return isinstance(obj, dict) and "label" in obj and "desc" in obj and isinstance(obj["desc"], str)

    def is_zero_with_unit(desc):
        if not isinstance(desc, str): return False
        desc = desc.replace("\xa0", " ")
        desc = re.sub(r"\s+", " ", desc).lower().strip()
        numbers = re.findall(r'\d*\.?\d+', desc)
        if not numbers: return False
        try:
            return all(float(n) == 0 for n in numbers)
        except:
            return False

    def clean_msrp_value(msrp):
        if msrp is None: return 0
        if isinstance(msrp, (int, float)): return msrp
        if isinstance(msrp, str):
            msrp = msrp.strip()
            try:
                value = float(msrp)
                return int(value) if value.is_integer() else value
            except ValueError:
                return 0
        return 0

    def clean_msrp_and_countries(model):
        general = model.get("general", {})
        general["msrp"] = clean_msrp_value(general.get("msrp"))
        countries = general.get("countries", [])
        if not isinstance(countries, list) or sorted(countries) != ["CA", "US"]:
            general["countries"] = ["US", "CA"]
        model["general"] = general
        return model

    def clean_specs_section(section, feature_list):
        if not isinstance(section, dict): return {}
        cleaned = {}
        for key, value in section.items():
            if not isinstance(value, dict): continue
            label = clean_string_global(value.get("label", ""))
            desc = value.get("desc", "")
            if not isinstance(desc, str): continue
            desc_clean = desc.strip().lower()

            if desc_clean == "yes":
                if label and label not in feature_list:
                    feature_list.append(label)
            elif desc_clean in REMOVE_VALUES or desc_clean in REMOVE_EXTRA or is_zero_with_unit(desc):
                continue
            else:
                cleaned[key] = value
        return cleaned

    def clean_specs_and_features(model):
        model.setdefault("general", {})
        model["general"]["msrp"] = clean_msrp_value(model["general"].get("msrp"))
        features = model.get("features", [])
        if not isinstance(features, list): features = []

        spec_headers = ["dimensions", "engine", "drivetrain", "operational", "hydraulics", "electrical", "weights", "measurements", "body", "other"]
        for header in spec_headers:
            if header in model:
                cleaned_section = clean_specs_section(model[header], features)
                if cleaned_section:
                    model[header] = cleaned_section
                else:
                    del model[header]
        model["features"] = features
        return model

    def remove_cross_duplicates(model):
        features = model.get("features", [])
        options = model.get("options", [])
        if not isinstance(features, list): features = []
        if not isinstance(options, list): options = []

        seen = set()
        unique_features = []
        for item in features:
            norm = normalize_text(item)
            if norm and norm not in seen:
                seen.add(norm)
                unique_features.append(item)

        unique_options = []
        for item in options:
            norm = normalize_text(item)
            if norm and norm not in seen:
                seen.add(norm)
                unique_options.append(item)

        model["features"] = unique_features
        model["options"] = unique_options
        return model

    def clean_meta_and_attachments(model):
        for k in ["_id", "updated_at", "created_at", "promotions"]:
            model.pop(k, None)
        general = model.get("general", {})
        if general.get("id") == 60733:
            del general["id"]
        meta = model.get("meta")
        if meta:
            if meta.get("source") == "CRS" or set(meta.keys()) == {"isAttachment", "isImplement"}:
                del model["meta"]
        return model

    def clean_whole_json(obj):
        if is_spec_object(obj):
            return {k: clean_dots_spec_string(v) if isinstance(v, str) else v for k, v in obj.items()}
        if isinstance(obj, dict):
            new_dict = {}
            for k, v in obj.items():
                cleaned_val = clean_whole_json(v)
                if is_blank_header(k, cleaned_val): continue
                new_dict[k] = cleaned_val
            return new_dict
        if isinstance(obj, list):
            return [clean_whole_json(x) for x in obj]
        if isinstance(obj, str):
            return clean_string_global(obj)
        return obj

    # --- UI GRAPHICS ---
    st.markdown("""
    <style>
    .block-container{ max-width:1200px; padding-top:2rem; }
    .main-title{ text-align:center; font-size:42px; font-weight:700; margin-bottom:10px; }
    .sub-title{ text-align:center; font-size:18px; color:#666; margin-bottom:30px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="main-title">🚀 Json Merged And Clean Data</div>
    <div class="sub-title">Upload files to merge (if >1), clean, parse MSRP, optimize, and standardize specs into clean camelCase properties</div>
    """, unsafe_allow_html=True)

    def parse_uploaded_files(uploaded_files):
        raw_data = []
        invalid_files = 0
        for uploaded_file in uploaded_files:
            try:
                if uploaded_file.name.lower().endswith(".json"):
                    data = json.load(uploaded_file)
                    if isinstance(data, list): raw_data.extend(data)
                    else: raw_data.append(data)
                elif uploaded_file.name.lower().endswith(".zip"):
                    with zipfile.ZipFile(uploaded_file, "r") as zip_ref:
                        for file_name in zip_ref.namelist():
                            if not file_name.lower().endswith(".json"): continue
                            try:
                                with zip_ref.open(file_name) as f:
                                    data = json.load(f)
                                    if isinstance(data, list): raw_data.extend(data)
                                    else: raw_data.append(data)
                            except:
                                invalid_files += 1
            except:
                invalid_files += 1
        return raw_data, invalid_files

    uploaded_files = st.file_uploader("📂 Upload JSON or ZIP Files", type=["json", "zip"], accept_multiple_files=True, key="studio_pro_uploader")

    if uploaded_files:
        start_time = time.time()
        total_files = len(uploaded_files)
        total_size = round(sum(file.size for file in uploaded_files) / (1024 * 1024), 2)
        
        progress = st.progress(0)
        raw_data, invalid_files = parse_uploaded_files(uploaded_files)
        progress.progress(0.3)

        # Merge check: Multiple files honge tabhi Merge Mode activate hoga
        is_merged_mode = total_files > 1

        cleaned_models = []
        for model in raw_data:
            model = clean_msrp_and_countries(model)
            model = clean_specs_and_features(model)
            model = remove_cross_duplicates(model)
            model = clean_meta_and_attachments(model)
            model = clean_whole_json(model)
            cleaned_models.append(model)
        
        progress.progress(0.6)
        cleaned_models = process_json_keys_to_camel(cleaned_models)
        progress.progress(0.8)
        
        # Multiple files me merge/sort hoga, Single file me original sequence (0, 1, 2...) exact preserve rahega
        if is_merged_mode:
            cleaned_models = sorted(
                cleaned_models,
                key=lambda x: (
                    str(x.get("general", {}).get("manufacturer", "")), 
                    str(x.get("general", {}).get("model", "")), 
                    str(x.get("general", {}).get("year", ""))
                )
            )
        
        progress.progress(1.0)
        elapsed = round(time.time() - start_time, 2)

        if is_merged_mode:
            st.success(f"🔥 Multi-file Merge & Optimization Complete! Processed {len(cleaned_models):,} structured records from {total_files} files.")
        else:
            st.success(f"🔥 Single JSON Cleaning Complete! Processed {len(cleaned_models):,} structured records in original sequence.")

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Files Uploaded", total_files)
        col2.metric("Processed Records", len(cleaned_models))
        col3.metric("Mode Active", "Merge + Clean" if is_merged_mode else "Clean (Sequence Preserved)")
        col4.metric("Total Size (MB)", total_size)
        st.caption(f"Engine Process Time: {elapsed} seconds")

        output_filename = "merged_cleaned_camelcase_oem_output.json" if is_merged_mode else "cleaned_camelcase_oem_output.json"

        json_output = json.dumps(cleaned_models, indent=2, ensure_ascii=False)
        st.download_button(
            label=f"⬇ Download {'Merged & ' if is_merged_mode else ''}Cleaned CamelCase JSON",
            data=json_output,
            file_name=output_filename,
            mime="application/json",
            use_container_width=True
        )

        with st.expander(f"👁️ Refactored JSON Preview ({min(len(cleaned_models),10)} of {len(cleaned_models)} records)"):
            preview_data = cleaned_models[:10]
            preview_json_string = json.dumps(preview_data, indent=2, ensure_ascii=False)
            st.code(preview_json_string, language="json")
    else:
        st.info("Upload one or more files to start the JSON cleaning and spec camelCase formatting engine.")

# =========================
# MODE 2: MODEL COMPARISON
# =========================
elif app_mode == "📊 Model Comparison":
    st.markdown("""
    <style>
    .block-container{ max-width:1200px; padding-top:2rem; }
    .main-title{ text-align:center; font-size:42px; font-weight:700; margin-bottom:10px; }
    .sub-title{ text-align:center; font-size:18px; color:#666; margin-bottom:30px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="main-title">📊 Model Comparison</div>
    <div class="sub-title">Upload your input Excel file containing 'OEM' and 'Prod' sheets to match models and generate final statuses</div>
    """, unsafe_allow_html=True)

    # --- HELPERS ---
    def tokenize_model(model_name):
        if pd.isna(model_name):
            return set()
        name = str(model_name).lower()
        name = name.replace("crew cab", "crewcab")
        name = name.replace("regular cab", "regularcab")
        name = name.replace("mega cab", "megacab")
        tokens = re.findall(r'[a-z]+|\d+x\d+|\d+', name)
        return set(tokens)

    uploaded_excel = st.file_uploader("📂 Upload Excel File (.xlsx)", type=["xlsx"], key="comparison_uploader")

    if uploaded_excel:
        try:
            with st.spinner("Processing Excel Sheets..."):
                oem_df = pd.read_excel(uploaded_excel, sheet_name="OEM")
                prod_df = pd.read_excel(uploaded_excel, sheet_name="Prod")

                for df in (oem_df, prod_df):
                    if 'URL' not in df.columns:
                        df['URL'] = ''
                    if 'Status' not in df.columns:
                        df['Status'] = pd.NA

                oem_df['Model_Tokens'] = oem_df['Model Name'].apply(tokenize_model)
                prod_df['Model_Tokens'] = prod_df['Model Name'].apply(tokenize_model)

                matched_rows = []
                matched_prod_indexes = set()

                for oem_idx, oem_row in oem_df.iterrows():
                    oem_tokens = oem_row['Model_Tokens']
                    match_found = False

                    for prod_idx, prod_row in prod_df.iterrows():
                        if prod_idx in matched_prod_indexes:
                            continue

                        prod_tokens = prod_row['Model_Tokens']

                        if oem_tokens == prod_tokens:
                            prod_df.at[prod_idx, 'Status'] = 'Exist'
                            oem_df.at[oem_idx, 'Status'] = 'Exist'
                            matched_prod_indexes.add(prod_idx)
                            match_found = True

                            matched_rows.append({
                                'OEM URL': oem_row['URL'],
                                'OEM Model': oem_row['Model Name'],
                                'Prod URL': prod_row['URL'],
                                'Prod Model': prod_row['Model Name']
                            })
                            break

                    if not match_found:
                        oem_df.at[oem_idx, 'Status'] = 'New'

                prod_df.loc[prod_df['Status'].isna(), 'Status'] = 'Discontinued'

                # Duplicate Handling
                prod_df['Token_Key'] = prod_df['Model_Tokens'].apply(lambda x: tuple(sorted(x)))
                dup_groups = prod_df.groupby('Token_Key').filter(lambda x: len(x) > 1)

                for _, group in dup_groups.groupby('Token_Key'):
                    first_idx = group.index[0]
                    prod_df.at[first_idx, 'Status'] = (
                        f"{prod_df.at[first_idx, 'Model Name']} "
                        f"({prod_df.at[first_idx, 'Status']})"
                    )
                    for idx in group.index[1:]:
                        prod_df.at[idx, 'Status'] = f"{prod_df.at[idx, 'Model Name']} (Duplicate)"

                matched_df = pd.DataFrame(matched_rows)

                uploaded_excel.seek(0)
                wb = load_workbook(uploaded_excel)

                def write_status(sheet_name, df):
                    sheet = wb[sheet_name]
                    sheet.cell(1, 3).value = 'Status'
                    for i, status in enumerate(df['Status'], start=2):
                        sheet.cell(i, 3).value = status

                write_status('Prod', prod_df)
                write_status('OEM', oem_df)

                if 'Matched Models' in wb.sheetnames:
                    del wb['Matched Models']

                match_sheet = wb.create_sheet('Matched Models')
                for r in dataframe_to_rows(matched_df, index=False, header=True):
                    match_sheet.append(r)

                output_stream = io.BytesIO()
                wb.save(output_stream)
                output_data = output_stream.getvalue()

                st.success("✅ Model comparison completed successfully!")

                base_name = os.path.splitext(uploaded_excel.name)[0]
                output_filename = f"{base_name}-latest.xlsx"

                st.download_button(
                    label="⬇ Download Processed Excel File",
                    data=output_data,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                tab1, tab2, tab3 = st.tabs(["OEM Sheet Preview", "Prod Sheet Preview", "Matched Models"])
                with tab1:
                    st.dataframe(oem_df[['Model Name', 'Status', 'URL']].head(50), use_container_width=True)
                with tab2:
                    st.dataframe(prod_df[['Model Name', 'Status', 'URL']].head(50), use_container_width=True)
                with tab3:
                    if not matched_df.empty:
                        st.dataframe(matched_df.head(50), use_container_width=True)
                    else:
                        st.info("No explicit matches found.")

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.warning("Please make sure your Excel sheet contains exactly 'OEM' and 'Prod' tab names with proper headers.")
    else:
        st.info("Awaiting Excel configuration upload to initialize token analysis.")
